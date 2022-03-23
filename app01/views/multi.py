from django.shortcuts import render, redirect

from app01 import models
from app01.utils.encrypt import md5


def multi_admin(request):
    """ 批量添加（Excel文件）"""
    from openpyxl import load_workbook
    if request.method == "GET":
        return render(request, 'multi_admin.html')
    file_object = request.FILES.get("exc")
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        username = row[0].value
        password = md5(row[1].value)
        exists = models.Admin.objects.filter(username=username).exists()
        if not exists:
            models.Admin.objects.create(username=username, password=password)
    return redirect('/admin/list/')


def multi_depart(request):
    """ 批量添加（Excel文件）"""
    from openpyxl import load_workbook
    if request.method == "GET":
        return render(request, 'multi_depart.html')
    file_object = request.FILES.get("exc")
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        title = row[0].value
        exists = models.Department.objects.filter(title=title).exists()
        if not exists:
            models.Department.objects.create(title=title)
    return redirect('/depart/list/')


def multi_pretty(request):
    """ 批量添加（Excel文件）"""
    from openpyxl import load_workbook
    if request.method == "GET":
        return render(request, 'multi_pretty.html')
    file_object = request.FILES.get("exc")
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        mobile = row[0].value
        price = row[1].value
        level = row[2].value
        status = row[3].value
        exists = models.PrettyNum.objects.filter(mobile=mobile).exists()
        if not exists:
            models.PrettyNum.objects.create(mobile=mobile, price=price, level=level, status=status)
    return redirect('/pretty/list/')


def multi_order(request):
    """ 批量添加（Excel文件）"""
    from openpyxl import load_workbook
    if request.method == "GET":
        return render(request, 'multi_order.html')
    file_object = request.FILES.get("exc")
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        oid = row[0].value
        title = row[1].value
        price = row[2].value
        status = row[3].value
        admin_id = row[4].value
        exists = models.Order.objects.filter(oid=oid).exists()
        if not exists:
            models.Order.objects.create(oid=oid, title=title, price=price, status=status, admin_id=admin_id)
    return redirect('/order/list/')


def multi_user(request):
    """ 批量添加（Excel文件）"""
    from openpyxl import load_workbook
    if request.method == "GET":
        return render(request, 'multi_user.html')
    file_object = request.FILES.get("exc")
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        name = row[0].value
        password = row[1].value
        age = row[2].value
        account = row[3].value
        create_time = row[4].value
        gender = row[5].value
        depart_id = row[6].value
        exists = models.UserInfo.objects.filter(name=name).exists()
        if not exists:
            models.UserInfo.objects.create(
                name=name,
                password=password,
                age=age,
                account=account,
                create_time=create_time,
                gender=gender,
                depart_id=depart_id)
    return redirect('/user/list/')
